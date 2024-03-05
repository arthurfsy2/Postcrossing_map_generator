from openpyxl import load_workbook
from urllib import parse, request
import json
import argparse
import os
import subprocess

# postcardStory_path = '../template/postcardStory.xlsx'
file_path = os.path.abspath(__file__)
dir_path = os.path.dirname(file_path)

postcardStory_path = os.path.abspath(os.path.join(
    dir_path, '..', '..', 'template', 'postcardStory.xlsx'))

repo_path = os.path.abspath(os.path.join(
    dir_path, '..', '..'))


def append_to_excel(card_id, content_original):
    wb = load_workbook(filename=postcardStory_path)
    sheet = wb['Sheet1']

    # 找到最后一行
    last_row = sheet.max_row

    # 遍历第一列，查找是否已经存在card_id
    found = False
    for row in range(1, last_row + 1):

        if sheet.cell(row=row, column=1).value == card_id:

            # 如果找到，就更新对应的内容
            sheet.cell(row=row, column=2, value=content_original)
            content_cn = translate(
                "da3f6df96672f9693e76ed14dd54f884", content_original, 'auto', 'zh')
            sheet.cell(row=row, column=3, value=content_cn)
            print(f"已存在明信片ID：{card_id}\n已替换原始内容：{
                  content_original}\n已替换翻译：{content_cn}")
            found = True
            break

    # 如果没有找到，就在最后一行追加数据
    if not found:
        sheet.cell(row=last_row+1, column=1, value=card_id)
        sheet.cell(row=last_row+1, column=2, value=content_original)
        content_cn = translate(
            "da3f6df96672f9693e76ed14dd54f884", content_original, 'auto', 'zh')
        sheet.cell(row=last_row+1, column=3, value=content_cn)
    # 保存文件
        print(f"已新增明信片ID：{card_id}\n原始内容：{content_original}\n翻译：{content_cn}")
    wb.save(postcardStory_path)


def translate(apikey, sentence, src_lan, tgt_lan):
    url = 'http://api.niutrans.com/NiuTransServer/translation?'
    data = {"from": src_lan, "to": tgt_lan, "apikey": apikey,
            "src_text": sentence.encode("utf-8")}
    data_en = parse.urlencode(data)
    req = url + "&" + data_en
    res = request.urlopen(req)
    res_dict = json.loads(res.read())
    if "tgt_text" in res_dict:
        result = res_dict['tgt_text']
    else:
        result = res
    return result


if __name__ == "__main__":
    parser = argparse.ArgumentParser()
    parser.add_argument("card_id", help="input your card_id")
    parser.add_argument("content_original", help="input your content_original")
    options = parser.parse_args()

    card_id = options.card_id
    content_original = options.content_original
    pull_command = ["git", "pull"]
    subprocess.run(pull_command, cwd=repo_path)
    print("已检查仓库更新")

    append_to_excel(card_id, content_original)
    # # 4. 提交仓库的修改，提交时备注为 tips 变量的值
    # tips = f"更新：{card_id}"
    # commit_command = ["git", "commit", "-am", f"{tips}"]
    # subprocess.run(commit_command, cwd=repo_path)

    # # 5. 推送修改到远程仓库
    # push_command = ["git", "push"]
    # subprocess.run(push_command, cwd=repo_path)
